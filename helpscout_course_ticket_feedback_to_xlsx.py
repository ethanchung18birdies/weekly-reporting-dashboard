#!/usr/bin/env python3
"""Add cleaned Help Scout customer feedback to the provided course-history workbook.

The default workbook already contains a ``Feedback Script Input`` sheet where
each course-related row has a Help Scout ``Ticket ID``. This script retrieves
those conversations, extracts customer-authored message bodies only, removes
HTML formatting, and writes the result to ``Script Collected Feedback``.

The source workbook is never changed. A new ``*_with_feedback.xlsx`` copy is
created beside it.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from copy import copy
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Dict, List, Optional
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode, urljoin
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


API_BASE = "https://api.helpscout.net/v2"
TOKEN_URL = f"{API_BASE}/oauth2/token"
DEFAULT_INPUT = Path.home() / "Downloads" / "18Birdies_Repeating_Courses_Type_History.xlsx"
DEFAULT_OUTPUT_DIR = Path.home() / "Desktop"
DEFAULT_SHEET = "Feedback Script Input"
TICKET_ID_HEADER = "Ticket ID"
FEEDBACK_HEADER = "Script Collected Feedback"
MAX_RETRIES = 6


class ConversationNotFoundError(RuntimeError):
    """Raised when a stored Help Scout conversation ID is no longer available."""


class HTMLTextExtractor(HTMLParser):
    """Convert Help Scout HTML to readable, paragraph-preserving text."""

    BLOCK_TAGS = {
        "address", "article", "aside", "blockquote", "br", "div", "dl", "dt", "dd",
        "fieldset", "figcaption", "figure", "footer", "form", "h1", "h2", "h3", "h4",
        "h5", "h6", "header", "hr", "li", "main", "nav", "ol", "p", "pre", "section",
        "table", "tr", "td", "th", "thead", "tbody", "tfoot", "ul",
    }

    def __init__(self) -> None:
        super().__init__()
        self.parts: List[str] = []
        self.skip_depth = 0

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag in {"script", "style"}:
            self.skip_depth += 1
        elif self.skip_depth == 0 and tag in self.BLOCK_TAGS:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style"} and self.skip_depth:
            self.skip_depth -= 1
        elif self.skip_depth == 0 and tag in self.BLOCK_TAGS:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self.skip_depth == 0 and data:
            self.parts.append(data)

    def get_text(self) -> str:
        text = unescape("".join(self.parts)).replace("\r", "\n")
        text = re.sub(r"[ \t]+\n", "\n", text)
        text = re.sub(r"\n[ \t]+", "\n", text)
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r"\n{3,}", "\n\n", text)
        return text.strip()


def strip_html(value: str) -> str:
    parser = HTMLTextExtractor()
    parser.feed(value or "")
    parser.close()
    return parser.get_text()


def render_progress(label: str, current: int, total: int) -> None:
    total = max(total, 1)
    current = min(current, total)
    width = 28
    filled = int(width * current / total)
    bar = "#" * filled + "-" * (width - filled)
    print(f"\r{label} [{bar}] {current}/{total}", end="", file=sys.stderr, flush=True)


def finish_progress() -> None:
    print(file=sys.stderr, flush=True)


def thread_is_customer(thread: dict) -> bool:
    created_by = thread.get("createdBy") or {}
    if str(created_by.get("type") or "").lower() == "customer":
        return True
    return bool(thread.get("customer"))


def extract_customer_feedback(conversation: dict) -> str:
    """Return all published customer messages, in ticket order, as plain text."""
    threads = conversation.get("_embedded", {}).get("threads", []) or []
    feedback_parts: List[str] = []

    for thread in sorted(threads, key=lambda item: item.get("createdAt") or ""):
        if str(thread.get("state") or "").lower() not in {"", "published"}:
            continue
        if str(thread.get("type") or "").lower() in {"note", "lineitem", "chatline"}:
            continue
        if not thread_is_customer(thread):
            continue

        raw = thread.get("body") or thread.get("plaintext") or thread.get("bodyPreview") or ""
        cleaned = strip_html(str(raw))
        if cleaned:
            feedback_parts.append(cleaned)

    return "\n\n---\n\n".join(feedback_parts)


class HelpScoutClient:
    def __init__(self, client_id: str, client_secret: str, pause_seconds: float) -> None:
        self.client_id = client_id
        self.client_secret = client_secret
        self.pause_seconds = pause_seconds
        self.access_token: Optional[str] = None

    def authenticate(self) -> None:
        payload = urlencode(
            {
                "grant_type": "client_credentials",
                "client_id": self.client_id,
                "client_secret": self.client_secret,
            }
        ).encode("utf-8")
        request = Request(
            TOKEN_URL,
            data=payload,
            headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"},
            method="POST",
        )
        try:
            with urlopen(request, timeout=30) as response:
                data = json.loads(response.read().decode("utf-8"))
        except HTTPError as exc:
            raise RuntimeError(f"Help Scout sign-in failed with HTTP {exc.code}: {exc.read().decode('utf-8', errors='replace')}") from exc
        except URLError as exc:
            raise RuntimeError(f"Could not reach Help Scout sign-in: {exc}") from exc

        self.access_token = data.get("access_token")
        if not self.access_token:
            raise RuntimeError("Help Scout did not return an access token.")

    def get_conversation(self, conversation_id: int) -> dict:
        if not self.access_token:
            raise RuntimeError("The Help Scout client is not authenticated.")

        url = urljoin(f"{API_BASE}/", f"conversations/{conversation_id}") + "?embed=threads"
        request = Request(url, headers={"Authorization": f"Bearer {self.access_token}", "Accept": "application/json"})
        last_error: Optional[Exception] = None

        for attempt in range(1, MAX_RETRIES + 1):
            try:
                with urlopen(request, timeout=60) as response:
                    data = json.loads(response.read().decode("utf-8"))
                if self.pause_seconds:
                    time.sleep(self.pause_seconds)
                return data
            except HTTPError as exc:
                if exc.code == 404:
                    raise ConversationNotFoundError(f"Ticket {conversation_id} is not available through the Help Scout API.") from exc
                if exc.code not in {429, 500, 502, 503, 504} or attempt == MAX_RETRIES:
                    body = exc.read().decode("utf-8", errors="replace")
                    raise RuntimeError(f"Help Scout API error {exc.code} for ticket {conversation_id}: {body}") from exc
                last_error = exc
                reason = f"HTTP {exc.code}"
            except (URLError, ConnectionResetError, TimeoutError) as exc:
                if attempt == MAX_RETRIES:
                    raise RuntimeError(f"Could not retrieve ticket {conversation_id}: {exc}") from exc
                last_error = exc
                reason = "network error"

            wait_seconds = min(30, 2 ** attempt)
            print(f"\nRetrying ticket {conversation_id} after {reason} in {wait_seconds}s...", file=sys.stderr)
            time.sleep(wait_seconds)

        raise RuntimeError(f"Exceeded retries for ticket {conversation_id}: {last_error}")


def header_columns(sheet) -> Dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def validate_workbook(input_path: Path, sheet_name: str) -> None:
    if load_workbook is None:
        raise RuntimeError("Missing 'openpyxl'. Install it with: python3 -m pip install openpyxl")
    if not input_path.is_file():
        raise RuntimeError(f"Workbook not found: {input_path}")
    if input_path.suffix.lower() != ".xlsx":
        raise RuntimeError("This script supports .xlsx workbooks only.")


def collect_ticket_rows(sheet, ticket_column: int, feedback_column: int, overwrite: bool) -> Dict[int, List[int]]:
    ticket_rows: Dict[int, List[int]] = {}
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, ticket_column).value
        existing_feedback = sheet.cell(row, feedback_column).value
        if value is None or (existing_feedback and not overwrite):
            continue
        try:
            ticket_id = int(str(value).strip())
        except (TypeError, ValueError):
            continue
        ticket_rows.setdefault(ticket_id, []).append(row)
    return ticket_rows


def fill_feedback(
    input_path: Path,
    output_path: Path,
    sheet_name: str,
    client: HelpScoutClient,
    overwrite: bool,
    checkpoint_every: int,
) -> tuple[int, int]:
    workbook = load_workbook(input_path)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' was not found. Available sheets: {', '.join(workbook.sheetnames)}")
    sheet = workbook[sheet_name]
    headers = header_columns(sheet)

    if TICKET_ID_HEADER not in headers:
        raise RuntimeError(f"Sheet '{sheet_name}' needs a '{TICKET_ID_HEADER}' column.")
    if FEEDBACK_HEADER not in headers:
        feedback_column = sheet.max_column + 1
        sheet.cell(1, feedback_column).value = FEEDBACK_HEADER
        sheet.cell(1, feedback_column)._style = sheet.cell(1, headers[TICKET_ID_HEADER])._style
        headers[FEEDBACK_HEADER] = feedback_column

    ticket_rows = collect_ticket_rows(sheet, headers[TICKET_ID_HEADER], headers[FEEDBACK_HEADER], overwrite)
    total = len(ticket_rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not total:
        print("No ticket rows need to be fetched. Use --overwrite to refresh existing feedback.")
        workbook.save(output_path)
        return 0, 0

    # Create a recoverable output immediately; subsequent checkpoints make
    # Ctrl-C safe and let a later run continue with already-filled rows.
    workbook.save(output_path)

    fetched = 0
    not_found = 0
    failures = 0
    try:
        for index, (ticket_id, rows) in enumerate(ticket_rows.items(), start=1):
            try:
                feedback = extract_customer_feedback(client.get_conversation(ticket_id))
                for row in rows:
                    cell = sheet.cell(row, headers[FEEDBACK_HEADER])
                    cell.value = feedback
                    alignment = copy(cell.alignment)
                    alignment.wrap_text = True
                    alignment.vertical = "top"
                    cell.alignment = alignment
                fetched += 1
            except ConversationNotFoundError:
                not_found += 1
            except RuntimeError as exc:
                failures += 1
                print(f"\nSkipping ticket {ticket_id}: {exc}", file=sys.stderr)

            if index % checkpoint_every == 0:
                workbook.save(output_path)
                print(f"\nCheckpoint saved after {index}/{total} tickets.", file=sys.stderr)
            render_progress("Fetching Help Scout tickets", index, total)
    except KeyboardInterrupt:
        workbook.save(output_path)
        finish_progress()
        print(f"Stopped safely. Saved progress to {output_path}.", file=sys.stderr)
        return fetched, failures + not_found
    finish_progress()

    workbook.save(output_path)
    print(f"{not_found} ticket IDs were not available through the Help Scout API.", file=sys.stderr)
    return fetched, failures + not_found


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Write cleaned Help Scout feedback into the course-history workbook.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help=f"Input .xlsx workbook (default: {DEFAULT_INPUT}).")
    parser.add_argument("--output", type=Path, help="Output .xlsx path. Defaults to the Desktop.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help=f"Worksheet to update (default: {DEFAULT_SHEET}).")
    parser.add_argument("--overwrite", action="store_true", help="Refresh rows that already have Script Collected Feedback.")
    parser.add_argument("--pause-seconds", type=float, default=0.1, help="Delay between Help Scout requests (default: 0.1).")
    parser.add_argument("--checkpoint-every", type=int, default=100, help="Save progress every N tickets (default: 100).")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else DEFAULT_OUTPUT_DIR / f"{input_path.stem}_with_feedback.xlsx"
    )
    validate_workbook(input_path, args.sheet)

    client_id = os.environ.get("HELPSCOUT_APP_ID") or os.environ.get("HELPSCOUT_CLIENT_ID")
    client_secret = os.environ.get("HELPSCOUT_APP_SECRET") or os.environ.get("HELPSCOUT_CLIENT_SECRET")
    if not client_id or not client_secret:
        print("Missing Help Scout credentials. Set HELPSCOUT_APP_ID and HELPSCOUT_APP_SECRET first.", file=sys.stderr)
        return 2

    client = HelpScoutClient(client_id, client_secret, args.pause_seconds)
    client.authenticate()
    if args.checkpoint_every < 1:
        print("--checkpoint-every must be at least 1.", file=sys.stderr)
        return 2
    fetched, failures = fill_feedback(
        input_path, output_path, args.sheet, client, args.overwrite, args.checkpoint_every
    )
    print(f"Wrote {fetched} ticket feedback entries to {output_path} ({failures} unavailable or skipped tickets).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
