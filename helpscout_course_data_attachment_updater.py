#!/usr/bin/env python3
"""Filter and extend the Course Entry Tickets workbook from Help Scout.

The script keeps rows whose customer feedback includes ``Reported clubId:``,
adds Course Name and Course ID columns parsed from that feedback, then appends
up to 250 newer, unique tickets assigned to Golf Course Scorecard Data that
both have an image attachment and include the same course-data marker.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from copy import copy
from datetime import datetime
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable, List, Optional
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode, urljoin
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


API_BASE = "https://api.helpscout.net/v2"
TOKEN_URL = f"{API_BASE}/oauth2/token"
DEFAULT_TEAM = "Golf Course Scorecard Data"
DEFAULT_APPEND_COUNT = 300
DEFAULT_INPUT = Path.home() / "Downloads" / "Source of Data Comparison Course Entry Tickets.xlsx"
DEFAULT_OUTPUT = Path.home() / "Downloads" / "Source of Data Comparison Course Entry Tickets - filtered and expanded.xlsx"
MAX_RETRIES = 6
IMAGE_EXTENSIONS = {".avif", ".bmp", ".gif", ".heic", ".heif", ".jpeg", ".jpg", ".png", ".tif", ".tiff", ".webp"}
COURSE_PATTERN = re.compile(r"Reported\s+clubId:\s*([^\s]+)\s+name:\s*(.*?)(?=\s*Feedback:)", re.IGNORECASE | re.DOTALL)


class HTMLTextExtractor(HTMLParser):
    BLOCK_TAGS = {"address", "article", "aside", "blockquote", "br", "div", "dl", "dt", "dd", "fieldset", "figcaption", "figure", "footer", "form", "h1", "h2", "h3", "h4", "h5", "h6", "header", "hr", "li", "main", "nav", "ol", "p", "pre", "section", "table", "tr", "td", "th", "thead", "tbody", "tfoot", "ul"}

    def __init__(self) -> None:
        super().__init__()
        self.parts: List[str] = []
        self.skip_depth = 0

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag in {"script", "style"}:
            self.skip_depth += 1
        elif self.skip_depth == 0 and tag in self.BLOCK_TAGS:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style"} and self.skip_depth:
            self.skip_depth -= 1
        elif self.skip_depth == 0 and tag in self.BLOCK_TAGS:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self.skip_depth == 0 and data:
            self.parts.append(data)

    def get_text(self) -> str:
        text = unescape("".join(self.parts)).replace("\r", "\n")
        text = re.sub(r"[ \t]+\n", "\n", text)
        text = re.sub(r"\n[ \t]+", "\n", text)
        text = re.sub(r"[ \t]+", " ", text)
        return re.sub(r"\n{3,}", "\n\n", text).strip()


def strip_html(value: str) -> str:
    parser = HTMLTextExtractor()
    parser.feed(value or "")
    parser.close()
    return parser.get_text()


def parse_course_details(feedback: str) -> tuple[str, str]:
    """Return Course Name and Course ID from the first reported club block."""
    match = COURSE_PATTERN.search(feedback or "")
    if not match:
        return "", ""
    course_id = match.group(1).strip()
    course_name = re.sub(r"\s+", " ", match.group(2)).strip()
    return course_name, course_id


def normalize_name(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def ticket_id_key(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def render_progress(label: str, current: int, total: int) -> None:
    total = max(total, 1)
    width = 28
    filled = int(width * min(current, total) / total)
    print(f"\r{label} [{'#' * filled}{'-' * (width - filled)}] {current}/{total}", end="", file=sys.stderr, flush=True)


def finish_progress() -> None:
    print(file=sys.stderr, flush=True)


def is_image_attachment(attachment: dict) -> bool:
    mime_type = str(attachment.get("mimeType") or "").lower()
    filename = str(attachment.get("filename") or attachment.get("fileName") or "").lower()
    return mime_type.startswith("image/") or any(filename.endswith(extension) for extension in IMAGE_EXTENSIONS)


def extract_images(threads: List[dict]) -> List[dict]:
    images: List[dict] = []
    for thread in threads:
        attachments = (thread.get("_embedded") or {}).get("attachments") or thread.get("attachments") or []
        for attachment in attachments:
            if not isinstance(attachment, dict) or not is_image_attachment(attachment):
                continue
            images.append(
                {
                    "filename": str(attachment.get("filename") or attachment.get("fileName") or "image"),
                    "download_url": str(((attachment.get("_links") or {}).get("download") or {}).get("href") or ""),
                }
            )
    return images


def thread_is_customer(thread: dict) -> bool:
    created_by = thread.get("createdBy") or {}
    return str(created_by.get("type") or "").lower() == "customer" or bool(thread.get("customer"))


def customer_feedback(threads: List[dict]) -> str:
    messages: List[str] = []
    for thread in sorted(threads, key=lambda item: item.get("createdAt") or ""):
        if not thread_is_customer(thread):
            continue
        if str(thread.get("state") or "").lower() not in {"", "published"}:
            continue
        if str(thread.get("type") or "").lower() in {"note", "lineitem", "chatline"}:
            continue
        text = strip_html(str(thread.get("body") or thread.get("plaintext") or thread.get("bodyPreview") or ""))
        if text:
            messages.append(text)
    return "\n\n---\n\n".join(messages)


def conversation_url(conversation: dict) -> str:
    href = ((conversation.get("_links") or {}).get("web") or {}).get("href")
    if href:
        return str(href)
    conversation_id, number = conversation.get("id"), conversation.get("number")
    return f"https://secure.helpscout.net/conversation/{conversation_id}/{number}/" if conversation_id and number else ""


class HelpScoutClient:
    def __init__(self, client_id: str, client_secret: str, pause_seconds: float) -> None:
        self.client_id = client_id
        self.client_secret = client_secret
        self.pause_seconds = pause_seconds
        self.access_token: Optional[str] = None

    def authenticate(self) -> None:
        payload = urlencode({"grant_type": "client_credentials", "client_id": self.client_id, "client_secret": self.client_secret}).encode("utf-8")
        request = Request(TOKEN_URL, data=payload, headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}, method="POST")
        try:
            with urlopen(request, timeout=30) as response:
                self.access_token = json.loads(response.read().decode("utf-8")).get("access_token")
        except (HTTPError, URLError) as exc:
            raise RuntimeError(f"Could not sign in to Help Scout: {exc}") from exc
        if not self.access_token:
            raise RuntimeError("Help Scout did not return an access token.")

    def request_json(self, path: str, **params) -> dict:
        if not self.access_token:
            raise RuntimeError("The Help Scout client is not authenticated.")
        url = urljoin(f"{API_BASE}/", path.lstrip("/"))
        query = urlencode({key: value for key, value in params.items() if value is not None})
        if query:
            url = f"{url}?{query}"
        request = Request(url, headers={"Authorization": f"Bearer {self.access_token}", "Accept": "application/json"})
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                with urlopen(request, timeout=60) as response:
                    payload = json.loads(response.read().decode("utf-8"))
                if self.pause_seconds:
                    time.sleep(self.pause_seconds)
                return payload
            except HTTPError as exc:
                if exc.code == 404:
                    return {}
                if exc.code not in {429, 500, 502, 503, 504} or attempt == MAX_RETRIES:
                    raise RuntimeError(f"Help Scout API error {exc.code} for {url}: {exc.read().decode('utf-8', errors='replace')}") from exc
                reason = f"HTTP {exc.code}"
            except (URLError, ConnectionResetError, TimeoutError) as exc:
                if attempt == MAX_RETRIES:
                    raise RuntimeError(f"Could not reach Help Scout API at {url}: {exc}") from exc
                reason = "network error"
            wait_seconds = min(30, 2 **attempt)
            print(f"\nRetrying after {reason} in {wait_seconds}s...", file=sys.stderr)
            time.sleep(wait_seconds)
        raise RuntimeError(f"Exceeded retries for {url}")

    def list_users(self) -> Iterable[dict]:
        page = 1
        while True:
            payload = self.request_json("/users", page=page, pageSize=100)
            users = payload.get("_embedded", {}).get("users", []) or []
            yield from users
            page_info = payload.get("page", {}) or {}
            if page >= (page_info.get("totalPages") or page_info.get("pages") or 1) or not users:
                break
            page += 1

    def list_recent_conversations(self, team_id: int, scan_limit: int) -> Iterable[dict]:
        page = 1
        yielded = 0
        while yielded < scan_limit:
            payload = self.request_json("/conversations", status="all", assigned_to=team_id, page=page, pageSize=100, sortField="createdAt", sortOrder="desc")
            conversations = payload.get("_embedded", {}).get("conversations", []) or []
            for conversation in conversations:
                yield conversation
                yielded += 1
                if yielded >= scan_limit:
                    return
            page_info = payload.get("page", {}) or {}
            if page >= (page_info.get("totalPages") or page_info.get("pages") or 1) or not conversations:
                break
            page += 1

    def get_conversation(self, conversation_id: int) -> dict:
        return self.request_json(f"/conversations/{conversation_id}")

    def get_threads(self, conversation_id: int) -> List[dict]:
        threads: List[dict] = []
        page = 1
        while True:
            payload = self.request_json(f"/conversations/{conversation_id}/threads", page=page, pageSize=100)
            page_threads = payload.get("_embedded", {}).get("threads", []) or []
            threads.extend(page_threads)
            page_info = payload.get("page", {}) or {}
            if page >= (page_info.get("totalPages") or page_info.get("pages") or 1) or not page_threads:
                break
            page += 1
        return threads


def user_display_name(user: dict) -> str:
    return str(user.get("name") or " ".join(part for part in (user.get("firstName") or "", user.get("lastName") or "") if part) or user.get("email") or "")


def resolve_team_id(client: HelpScoutClient, team_name: str) -> int:
    matches = [user for user in client.list_users() if normalize_name(user_display_name(user)) == normalize_name(team_name)]
    if len(matches) == 1:
        return int(matches[0]["id"])
    if not matches:
        raise RuntimeError(f"Could not find a Help Scout team/user named '{team_name}'.")
    raise RuntimeError(f"More than one Help Scout team/user matched '{team_name}'. Use --team-id to specify one.")


def parse_created_at(value: str):
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
    except (TypeError, ValueError):
        return value


def validate_input(input_path: Path) -> None:
    if load_workbook is None:
        raise RuntimeError("Missing openpyxl. Run this script with your helpscout-venv environment.")
    if not input_path.is_file():
        raise RuntimeError(f"Workbook not found: {input_path}")


def filter_and_append(input_path: Path, output_path: Path, client: HelpScoutClient, team_id: int, append_count: int, scan_limit: int) -> tuple[int, int, int]:
    workbook = load_workbook(input_path)
    if "Tickets" not in workbook.sheetnames:
        raise RuntimeError("Expected a worksheet named 'Tickets'.")
    sheet = workbook["Tickets"]
    headers = {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value}
    required = {"ticket_id", "helpscout_url", "created_at", "customer_feedback", "image_count", "image_filenames", "image_download_links"}
    missing = required - set(headers)
    if missing:
        raise RuntimeError(f"Missing expected workbook columns: {', '.join(sorted(missing))}")

    existing_ticket_ids = {ticket_id_key(sheet.cell(row, headers["ticket_id"]).value) for row in range(2, sheet.max_row + 1)}
    feedback_column = headers["customer_feedback"]
    removed = 0
    for row in range(sheet.max_row, 1, -1):
        if "Reported clubId:" not in str(sheet.cell(row, feedback_column).value or ""):
            sheet.delete_rows(row, 1)
            removed += 1

    course_name_column = headers.get("Course Name") or sheet.max_column + 1
    if "Course Name" not in headers:
        sheet.cell(1, course_name_column).value = "Course Name"
        sheet.cell(1, course_name_column)._style = copy(sheet.cell(1, headers["customer_feedback"])._style)
    course_id_column = headers.get("Course ID") or course_name_column + 1
    if "Course ID" not in headers:
        sheet.cell(1, course_id_column).value = "Course ID"
        sheet.cell(1, course_id_column)._style = copy(sheet.cell(1, headers["ticket_id"])._style)
    sheet.column_dimensions[sheet.cell(1, course_name_column).column_letter].width = 42
    sheet.column_dimensions[sheet.cell(1, course_id_column).column_letter].width = 38

    for row in range(2, sheet.max_row + 1):
        course_name, course_id = parse_course_details(str(sheet.cell(row, feedback_column).value or ""))
        sheet.cell(row, course_name_column).value = course_name
        sheet.cell(row, course_id_column).value = course_id

    # Reuse a current data-row style so appended API records match the sheet.
    template_row = 2
    appended = 0
    inspected = 0
    for summary in client.list_recent_conversations(team_id, scan_limit):
        if appended >= append_count:
            break
        conversation_id = summary.get("id")
        if not conversation_id:
            continue
        inspected += 1
        conversation = client.get_conversation(int(conversation_id))
        if not conversation:
            continue
        ticket_id = ticket_id_key(conversation.get("number") or conversation.get("id"))
        if not ticket_id or ticket_id in existing_ticket_ids:
            continue
        threads = client.get_threads(int(conversation_id))
        images = extract_images(threads)
        if not images:
            continue
        feedback = customer_feedback(threads)
        if "Reported clubId:" not in feedback:
            continue
        course_name, course_id = parse_course_details(feedback)
        new_row = sheet.max_row + 1
        values = [
            ticket_id,
            conversation_url(conversation),
            parse_created_at(str(conversation.get("createdAt") or "")),
            feedback,
            len(images),
            " | ".join(image["filename"] for image in images),
            " | ".join(image["download_url"] for image in images if image["download_url"]),
            course_name,
            course_id,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(new_row, column)
            cell.value = value
            style_column = headers["customer_feedback"] if column == course_name_column else headers["ticket_id"] if column == course_id_column else column
            cell._style = copy(sheet.cell(template_row, style_column)._style)
            cell.alignment = copy(sheet.cell(template_row, style_column).alignment)
        appended += 1
        existing_ticket_ids.add(ticket_id)
        render_progress("Finding qualifying tickets", appended, append_count)
    if append_count:
        finish_progress()

    sheet.auto_filter.ref = f"A1:{sheet.cell(1, course_id_column).column_letter}{sheet.max_row}"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return removed, appended, inspected


def main() -> int:
    parser = argparse.ArgumentParser(description="Filter and extend a course-data ticket workbook with Help Scout records.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help=f"Input workbook (default: {DEFAULT_INPUT}).")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help=f"Output workbook (default: {DEFAULT_OUTPUT}).")
    parser.add_argument("--team", default=DEFAULT_TEAM, help=f"Help Scout team/user name (default: {DEFAULT_TEAM}).")
    parser.add_argument("--team-id", type=int, help="Help Scout team/user ID; overrides --team.")
    parser.add_argument("--append-count", type=int, default=DEFAULT_APPEND_COUNT, help="New qualifying tickets to append (default: 300).")
    parser.add_argument("--scan-limit", type=int, default=5000, help="Maximum recent team tickets to inspect (default: 5000).")
    parser.add_argument("--pause-seconds", type=float, default=0.1, help="Delay between API requests (default: 0.1).")
    args = parser.parse_args()
    if args.append_count < 0 or args.scan_limit < 1:
        print("--append-count must be zero or more and --scan-limit must be at least 1.", file=sys.stderr)
        return 2

    input_path, output_path = args.input.expanduser().resolve(), args.output.expanduser().resolve()
    validate_input(input_path)
    client_id, client_secret = os.environ.get("HELPSCOUT_CLIENT_ID"), os.environ.get("HELPSCOUT_CLIENT_SECRET")
    if not client_id or not client_secret:
        print("Missing Help Scout credentials. Set HELPSCOUT_CLIENT_ID and HELPSCOUT_CLIENT_SECRET first.", file=sys.stderr)
        return 2
    client = HelpScoutClient(client_id, client_secret, args.pause_seconds)
    client.authenticate()
    team_id = args.team_id or resolve_team_id(client, args.team)
    removed, appended, inspected = filter_and_append(input_path, output_path, client, team_id, args.append_count, args.scan_limit)
    print(f"Removed {removed} rows, appended {appended} new qualifying tickets after inspecting {inspected}, and wrote {output_path}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
